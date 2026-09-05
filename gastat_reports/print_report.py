# Copyright (c) 2026, GASTAT and contributors
# For license information, please see license.txt

"""PDF and Excel export for GASTAT reports."""

import base64
import io
import os

import frappe
from frappe.utils import flt, now_datetime, cint

from gastat_reports.api import get_production_report, get_employee_statistics, _log_report
from gastat_reports.utils import get_month_name_ar


# ============================================================================
# Shared HTML helpers
# ============================================================================

CSS = """
@page { size: A4; margin: 0; }
* { box-sizing: border-box; }
html, body {
    margin: 0; padding: 0;
    font-family: 'Segoe UI', Tahoma, Arial, sans-serif;
    color: #1f2937;
    direction: rtl;
    background: #ffffff;
}
.page {
    position: relative;
    min-height: 1080px;
    padding: 14mm 12mm 22mm 12mm;
}
/* Letterhead */
.letterhead {
    display: flex; align-items: center; justify-content: space-between;
    border-bottom: 3px solid #0f766e;
    padding-bottom: 6mm; width: 100%;
}
.letterhead .logo img { max-height: 34mm; max-width: 60mm; }
.letterhead .title-block { text-align: center; flex: 1; }
.letterhead .title-block h1 {
    margin: 0; font-size: 15pt; color: #0f766e;
    font-weight: 700;
}
.letterhead .title-block .subtitle {
    margin: 2px 0 0 0; font-size: 10pt; color: #475569;
}
.letterhead .spacer { width: 60mm; }
.letterhead .spacer-sm { width: 20mm; }

.header-meta {
    display: flex; justify-content: space-between;
    margin-top: 4mm; font-size: 9.5pt; color: #374151;
    border: 1px solid #e2e8f0; border-radius: 6px;
    padding: 3mm 4mm; background: #f8fafc;
}
.header-meta div strong { color: #0f766e; }

h2.section-title {
    font-size: 13pt; color: #0f766e; margin: 6mm 0 3mm;
    border-right: 4px solid #14b8a6; padding-right: 3mm;
}

table.data {
    width: 100%; border-collapse: collapse; font-size: 9pt;
}
table.data th {
    background: #0f766e; color: #ffffff; padding: 2.2mm 2mm;
    font-weight: 600; border: 1px solid #0f766e; text-align: center;
}
table.data td {
    padding: 2mm 2mm; border: 1px solid #cbd5e1; text-align: center;
}
table.data tbody tr:nth-child(even) { background: #f1f5f9; }
table.data tbody tr.total td {
    background: #0f766e !important; color: #ffffff; font-weight: 700;
}
td.num, th.num { text-align: left; }
td.r, th.r { text-align: right; }

.summary-box {
    display: flex; flex-wrap: wrap; gap: 3mm;
    margin: 4mm 0;
}
.summary-box .card {
    flex: 1 1 22%; border: 1px solid #e2e8f0; border-radius: 8px;
    padding: 3mm; text-align: center; background: #f8fafc;
}
.summary-box .card .k { font-size: 8.5pt; color: #64748b; margin-bottom: 1mm; }
.summary-box .card .v { font-size: 12pt; font-weight: 700; color: #0f766e; }

.notes {
    margin-top: 5mm; font-size: 8.5pt; color: #64748b;
    border-top: 1px solid #e2e8f0; padding-top: 2mm;
}
.signature-block {
    display: flex; justify-content: space-between; align-items: flex-end;
    margin-top: 16mm;
}
.signature-block .sig {
    text-align: center; width: 45%;
}
.signature-block .sig .line { border-bottom: 1px solid #334155; margin-bottom: 2mm; }
.signature-block .sig .name { font-weight: 700; }
.signature-block .sig .title { font-size: 9pt; color: #475569; }
.footer-note {
    position: absolute; bottom: 8mm; left: 12mm; right: 12mm;
    text-align: center; font-size: 8pt; color: #94a3b8;
    border-top: 1px solid #e2e8f0; padding-top: 2mm;
}
.employee-section-header td {
    background: #e2e8f0 !important; color: #0f766e;
    font-weight: 700; font-size: 10pt;
}
"""


def _img_data_uri(file_url, fallback_text=""):
	"""Convert an attachment URL to a base64 data URI, or return empty."""
	if not file_url:
		return ""
	try:
		path = file_url
		if file_url.startswith("/files/"):
			path = os.path.join(frappe.local.conf.get("sites_path"), "sites", frappe.local.site,
				file_url.lstrip("/"))
		elif file_url.startswith("/private"):
			path = os.path.join(frappe.local.conf.get("sites_path"), "sites", frappe.local.site,
				file_url.lstrip("/"))
		elif file_url.startswith("http"):
			return file_url
		if not os.path.exists(path):
			frappe.log_error(f"Logo file not found at {path}")
			return ""
		with open(path, "rb") as fh:
			data = base64.b64encode(fh.read()).decode()
		ext = os.path.splitext(path)[1].lstrip(".").lower()
		mime = {"png": "image/png", "jpg": "image/jpeg", "jpeg": "image/jpeg", "gif": "image/gif", "svg": "image/svg+xml"}.get(ext, "image/png")
		return f"data:{mime};base64,{data}"
	except Exception as e:
		frappe.log_error(f"Logo embed failed: {e}")
		return ""


def _monetary(value):
	return f"{flt(value):,.2f}"


def _build_html(letterhead_html, body_html, footer_text, page_title):
	settings = frappe.get_single("GASTAT Settings")
	return f"""
<!DOCTYPE html>
<html lang="ar" dir="rtl">
<head>
<meta charset="utf-8"/>
<title>{page_title}</title>
<style>{CSS}</style>
</head>
<body>
<div class="page">
    {letterhead_html}
    {body_html}
    <div class="footer-note">{footer_text or '&nbsp;'}</div>
</div>
</body>
</html>
"""


def _letterhead(title, subtitle):
	settings = frappe.get_single("GASTAT Settings")
	logo_url = settings.report_header_logo
	logo_img = ""
	if logo_url:
		uri = _img_data_uri(logo_url)
		if uri:
			logo_img = f'<img src="{uri}"/>'
	return f"""
<div class="letterhead">
    <div class="logo">{logo_img}</div>
    <div class="title-block">
        <h1>{title}</h1>
        <div class="subtitle">{subtitle}</div>
    </div>
    <div class="spacer"></div>
</div>
"""


def _header_meta(summary):
	company_name = summary.get("company_name") or summary.get("company") or "-"
	month_ar = get_month_name_ar(summary.get("month", 1), summary.get("year"))
	period = f"{month_ar} {summary.get('year')}"
	generated = frappe.utils.format_datetime(now_datetime(), "dd MMMM yyyy - hh:mm a")
	return f"""
<div class="header-meta">
    <div><strong>الشركة:</strong> {company_name}</div>
    <div><strong>الفترة:</strong> {period}</div>
    <div><strong>تاريخ التوليد:</strong> {generated}</div>
</div>
"""


def _signature_cells(cell, cell2=None):
	return f"""
<div class="signature-block">
    <div class="sig">
        <div class="line">&nbsp;</div>
        <div class="name">المعتمد / Authorized Signatory</div>
        <div class="title">{cell}</div>
    </div>
</div>
"""


def _build_pdf(html):
	options = {
		"page-size": "A4",
		"margin-top": "0mm",
		"margin-bottom": "0mm",
		"margin-left": "0mm",
		"margin-right": "0mm",
		"encoding": "UTF-8",
		"enable-local-file-access": None,
	}
	from frappe.utils.pdf import get_pdf
	return get_pdf(html, options=options)


# ============================================================================
# Production HTML
# ============================================================================

def production_html(data):
	summary = data["summary"]
	rows = data["rows"]
	settings = frappe.get_single("GASTAT Settings")

	title = "تقرير المسح الصناعي الشهري"
	subtitle = "Industrial Production Survey – Monthly Report"
	head = _letterhead(title, subtitle)
	meta = _header_meta(summary)

	# summary cards
	ret_card = f"""
        <div class="card"><div class="k">خصم المرتجعات</div><div class="v">{_monetary(summary.get('returns_value', 0.0))} {summary.get('currency','')}</div></div>""" if flt(summary.get("returns_value", 0.0)) else ""
	cards = f"""
    <div class="summary-box">
        <div class="card"><div class="k">عدد الأصناف المنتجة</div><div class="v">{summary['total_items']}</div></div>
        <div class="card"><div class="k">إجمالي الكمية</div><div class="v">{_monetary(summary['total_qty'])}</div></div>
        <div class="card"><div class="k">متوسط سعر الوحدة</div><div class="v">{_monetary(summary['avg_unit_price'])}</div></div>
        <div class="card"><div class="k">إجمالي قيمة الإنتاج (الصافي)</div><div class="v">{_monetary(summary['total_value'])} {summary.get('currency','')}</div></div>
        {ret_card}
    </div>
    """

	trs = ""
	for r in rows:
		trs += f"""
        <tr>
            <td>{r['sr']}</td>
            <td class="r">{r['item_code']}</td>
            <td class="r">{r['item_name']}</td>
            <td>{r['hs_code'] or '-'}</td>
            <td>{_monetary(r['qty'])}</td>
            <td>{r['uom']}</td>
            <td class="num">{_monetary(r['unit_price'])}</td>
            <td class="num">{_monetary(r['total_value'])}</td>
        </tr>"""
	grand_qty = sum(r['qty'] for r in rows)
	grand_val = sum(r['total_value'] for r in rows)
	ret_qty = flt(summary.get('returns_qty', 0.0))
	ret_val = flt(summary.get('returns_value', 0.0))
	ret_rows = f"""
        <tr>
            <td colspan="4" class="r">خصم المرتجعات / Less: Returns</td>
            <td>{_monetary(ret_qty)}</td><td></td><td></td>
            <td class="num">{_monetary(ret_val)}</td>
        </tr>""" if ret_val else ""
	net_rows = f"""
        <tr class="total">
            <td colspan="4">الصافي / Net Total</td>
            <td>{_monetary(grand_qty + ret_qty)}</td><td></td><td></td>
            <td class="num">{_monetary(summary['total_value'])} {summary.get('currency','')}</td>
        </tr>"""
	table = f"""
    <h2 class="section-title">تفاصيل الإنتاج حسب الصنف</h2>
    <table class="data">
        <thead>
            <tr>
                <th>#</th><th>كود الصنف</th><th>اسم الصنف</th><th>رمز HS</th>
                <th>الكمية المنتجة</th><th>وحدة القياس</th><th>سعر الوحدة</th><th>القيمة الإجمالية</th>
            </tr>
        </thead>
        <tbody>
            {trs}
            <tr class="total">
                <td colspan="4">إجمالي المبيعات / Gross Total</td>
                <td>{_monetary(grand_qty)}</td><td></td><td></td>
                <td class="num">{_monetary(grand_val)} {summary.get('currency','')}</td>
            </tr>
            {ret_rows}
            {net_rows}
        </tbody>
    </table>
    """

	# signature
	auth_name = settings.authorized_signatory_name or ""
	auth_title = settings.authorized_signatory_title or ""
	sig = ""
	if auth_name:
		sig += f"<div class='name'>{auth_name}</div>"
	if auth_title:
		sig += f"<div class='title'>{auth_title}</div>"
	else:
		sig = ""

	signature = f"""
    <div class="signature-block">
        <div class="sig">
            <div class="line">&nbsp;</div>
            <div class="name">المعتمد / Authorized Signatory</div>
            <div class="title">{sig or '&nbsp;'}</div>
        </div>
    </div>
    """

	notes = f"<div class='notes'>مصدر البيانات: {summary.get('production_source','')} | مصدر الأسعار: {summary.get('price_source','')} | العملة: {summary.get('currency','')} | ملاحظات: {settings.report_footer_text or ''}</div>"

	body = meta + cards + table + notes + signature
	return _build_html(head, body, settings.report_footer_text, title)


# ============================================================================
# Employee HTML
# ============================================================================

def _salary_table(title, rows, currency, footer_label):
	trs = ""
	total = 0.0
	for i, r in enumerate(rows, start=1):
		total += r["monthly_salary"]
		gender = "ذكر" if r["gender"] == "Male" else ("أنثى" if r["gender"] == "Female" else "-")
		trs += f"""
        <tr>
            <td>{i}</td>
            <td class="r">{r['employee']}</td>
            <td class="r">{r['employee_name']}</td>
            <td>{r['national_id'] or '-'}</td>
            <td class="r">{r['nationality'] or '-'}</td>
            <td>{gender}</td>
            <td>{r['category']}</td>
            <td class="num">{_monetary(r['monthly_salary'])}</td>
        </tr>"""
	table = f"""
    <h2 class="section-title">{title} ({len(rows)})</h2>
    <table class="data">
        <thead>
            <tr>
                <th>#</th><th>رقم الموظف</th><th>اسم الموظف</th><th>الهوية / الإقامة</th>
                <th>الجنسية</th><th>الجنس</th><th>الفئة</th><th>الراتب الشهري</th>
            </tr>
        </thead>
        <tbody>
            {trs}
            <tr class="total">
                <td colspan="7">{footer_label}</td>
                <td class="num">{_monetary(total)} {currency}</td>
            </tr>
        </tbody>
    </table>
    """
	return table


def employee_html(data):
	summary = data["summary"]
	saudi_rows = data["saudi_rows"]
	nonsaudi_rows = data["nonsaudi_rows"]
	settings = frappe.get_single("GASTAT Settings")

	title = "تقرير إحصاءات الموظفين"
	subtitle = "Employee Statistics – Monthly Report"
	head = _letterhead(title, subtitle)
	meta = _header_meta(summary)
	currency = summary.get("currency", "")

	cards = f"""
    <div class="summary-box">
        <div class="card"><div class="k">نسبة السعوديين</div><div class="v">{summary['saudi_pct']}%</div></div>
        <div class="card"><div class="k">نسبة غير السعوديين</div><div class="v">{summary['nonsaudi_pct']}%</div></div>
        <div class="card"><div class="k">سعوديون (ذكور)</div><div class="v">{summary['saudi_male']}</div></div>
        <div class="card"><div class="k">سعوديات (إناث)</div><div class="v">{summary['saudi_female']}</div></div>
        <div class="card"><div class="k">غير سعوديين (ذكور)</div><div class="v">{summary['nonsaudi_male']}</div></div>
        <div class="card"><div class="k">غير سعوديات (إناث)</div><div class="v">{summary['nonsaudi_female']}</div></div>
        <div class="card"><div class="k">إجمالي رواتب السعوديين</div><div class="v">{_monetary(summary['saudi_salaries'])} {currency}</div></div>
        <div class="card"><div class="k">إجمالي رواتب غير السعوديين</div><div class="v">{_monetary(summary['nonsaudi_salaries'])} {currency}</div></div>
    </div>
    """

	table_saudi = _salary_table("الموظفون السعوديون", saudi_rows, currency, "إجمالي رواتب السعوديين")
	table_nonsaudi = _salary_table("الموظفون غير السعوديين", nonsaudi_rows, currency, "إجمالي رواتب غير السعوديين")

	auth_name = settings.authorized_signatory_name or ""
	auth_title = settings.authorized_signatory_title or ""
	sig_inner = (f"<div class='name'>{auth_name}</div>" if auth_name else "") + (f"<div class='title'>{auth_title}</div>" if auth_title else "")
	signature = f"""
    <div class="signature-block">
        <div class="sig">
            <div class="line">&nbsp;</div>
            <div class="name">المعتمد / Authorized Signatory</div>
            <div class="title">{sig_inner or '&nbsp;'}</div>
        </div>
    </div>
    """

	notes = f"<div class='notes'>مرجع الراتب: {summary.get('salary_col','')} | العملة: {currency} | ملاحظات: {settings.report_footer_text or ''}</div>"

	body = meta + cards + table_saudi + table_nonsaudi + notes + signature
	return _build_html(head, body, settings.report_footer_text, title)


# ============================================================================
# Whitelisted export endpoints
# ============================================================================

def _save_and_return(html, filename, report_type, month_no, year, company):
	pdf_data = _build_pdf(html)
	file_doc = frappe.get_doc({
		"doctype": "File",
		"file_name": filename,
		"is_private": 1,
		"content": pdf_data,
	})
	file_doc.save(ignore_permissions=True)
	# audit log
	name = _log_report(report_type, month_no, year, company, file_doc.file_url)
	frappe.db.commit()
	return file_doc.file_url


@frappe.whitelist()
def export_production_pdf(company=None, month=0, year=0, **kwargs):
	data = get_production_report(company=company, month=month or 0, year=year or 0)
	html = production_html(data)
	summary = data["summary"]
	filename = f"Production_Survey_{summary['month']:02d}_{summary['year']}.pdf"
	return _save_and_return(html, filename, "Production", summary["month"], summary["year"], summary["company"])


@frappe.whitelist()
def export_employee_pdf(company=None, month=0, year=0, **kwargs):
	data = get_employee_statistics(company=company, month=month or 0, year=year or 0)
	html = employee_html(data)
	summary = data["summary"]
	filename = f"Employee_Statistics_{summary['month']:02d}_{summary['year']}.pdf"
	return _save_and_return(html, filename, "Employees", summary["month"], summary["year"], summary["company"])


# ============================================================================
# Excel exports (openpyxl)
# ============================================================================

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


_TEAL = "0F766E"
_LIGHT = "E2E8F0"
_TOTAL = "B45309"


def _xlsx_style():
	thin = Side(style="thin", color="CBD5E1")
	return {
		"header_fill": PatternFill("solid", fgColor=_TEAL),
		"header_font": Font(color="FFFFFF", bold=True, name="Calibri", size=11),
		"sub_fill": PatternFill("solid", fgColor=_LIGHT),
		"sub_font": Font(color=_TEAL, bold=True, size=11),
		"total_fill": PatternFill("solid", fgColor=_TOTAL),
		"total_font": Font(color="FFFFFF", bold=True, size=11),
		"border": Border(left=thin, right=thin, top=thin, bottom=thin),
		"center": Alignment(horizontal="center", vertical="center"),
		"right": Alignment(horizontal="right", vertical="center"),
	}


def _write_header(ws, title, subtitle, ncols, style):
	ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
	c = ws.cell(row=1, column=1, value=title)
	c.font = Font(bold=True, size=16, color=_TEAL)
	c.alignment = style["center"]
	ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
	c2 = ws.cell(row=2, column=1, value=subtitle)
	c2.font = Font(italic=True, size=11, color="64748B")
	c2.alignment = style["center"]
	ws.row_dimensions[1].height = 28
	ws.row_dimensions[2].height = 16


def _write_table_header(ws, row_idx, headers, style):
	for col, h in enumerate(headers, start=1):
		c = ws.cell(row=row_idx, column=col, value=h)
		c.fill = style["header_fill"]
		c.font = style["header_font"]
		c.alignment = style["center"]
		c.border = style["border"]
	ws.row_dimensions[row_idx].height = 20
	return row_idx + 1


def _autofit(ws, widths):
	for i, w in enumerate(widths, start=1):
		ws.column_dimensions[get_column_letter(i)].width = w


def _save_xlsx(wb, filename, report_type, month_no, year, company):
	buf = io.BytesIO()
	wb.save(buf)
	buf.seek(0)
	file_doc = frappe.get_doc({
		"doctype": "File",
		"file_name": filename,
		"is_private": 1,
		"content": buf.getvalue(),
	})
	file_doc.save(ignore_permissions=True)
	name = _log_report(report_type, month_no, year, company, file_doc.file_url)
	frappe.db.commit()
	return file_doc.file_url


@frappe.whitelist()
def export_production_excel(company=None, month=0, year=0, **kwargs):
	data = get_production_report(company=company, month=month or 0, year=year or 0)
	summary = data["summary"]
	rows = data["rows"]
	style = _xlsx_style()
	wb = Workbook()
	ws = wb.active
	ws.title = "Production Survey"

	title = "تقرير المسح الصناعي الشهري (Industrial Production Survey)"
	subtitle = (f"{summary['company_name']} | {summary['month_name_ar']} {summary['year']}"
	            f" | Data source: {summary.get('production_source','')} | Currency: {summary.get('currency','')} | Price source: {summary.get('price_source','')}")
	_write_header(ws, title, subtitle, 8, style)

	# summary block
	srow = 4
	summary_items = [
		("Total Items Produced", summary["total_items"]),
		("Total Quantity", flt(summary["total_qty"])),
		("Average Unit Price", flt(summary["avg_unit_price"])),
		("Total Production Value", flt(summary["total_value"])),
	]
	for i, (k, v) in enumerate(summary_items):
		c = ws.cell(row=srow, column=1 + i * 2, value=k)
		c.font = Font(bold=True, size=10)
		c2 = ws.cell(row=srow, column=2 + i * 2, value=flt(v))
		c2.font = Font(bold=True, size=10, color=_TEAL)
	srow += 3

	headers = ["Row #", "Item Code", "Item Name", "HS Code", "Quantity Produced", "UOM", "Unit Price", "Total Value"]
	r = _write_table_header(ws, srow, headers, style)
	grand_qty = 0.0
	grand_val = 0.0
	for row in rows:
		grand_qty += row["qty"]
		grand_val += row["total_value"]
		vals = [row["sr"], row["item_code"], row["item_name"], row["hs_code"] or "-",
		        flt(row["qty"]), row["uom"], flt(row["unit_price"]), flt(row["total_value"])]
		for col, v in enumerate(vals, start=1):
			c = ws.cell(row=r, column=col, value=v)
			c.border = style["border"]
			c.alignment = style["right"] if col in (2, 3) else style["center"]
		r += 1

	# row writers
	def _total_row(label, qty, val, fill, font):
		nonlocal r
		vals = ["", "", label, "", flt(qty), "", "", flt(val)]
		for col, v in enumerate(vals, start=1):
			c = ws.cell(row=r, column=col, value=v)
			c.fill = fill
			c.font = font
			c.border = style["border"]
			c.alignment = style["center"]
		r += 1

	_total_row("Gross Total", grand_qty, grand_val, style["total_fill"], style["total_font"])
	ret_val = flt(summary.get("returns_value", 0.0))
	if ret_val:
		ret_fill = PatternFill("solid", fgColor="FFF7ED")
		ret_font = Font(color="C2410C", bold=True, size=11)
		_total_row("Less: Returns", flt(summary.get("returns_qty", 0.0)), ret_val, ret_fill, ret_font)
	_total_row("Net Total", flt(summary["total_qty"]), flt(summary["total_value"]),
	           style["total_fill"], Font(color="FFFFFF", bold=True, size=12))

	_autofit(ws, [8, 14, 30, 14, 16, 10, 14, 16])
	ws.freeze_panes = f"A{srow+2}"
	filename = f"Production_Survey_{summary['month']:02d}_{summary['year']}.xlsx"
	return _save_xlsx(wb, filename, "Production", summary["month"], summary["year"], summary["company"])


@frappe.whitelist()
def export_employee_excel(company=None, month=0, year=0, **kwargs):
	data = get_employee_statistics(company=company, month=month or 0, year=year or 0)
	summary = data["summary"]
	saudi_rows = data["saudi_rows"]
	nonsaudi_rows = data["nonsaudi_rows"]
	style = _xlsx_style()
	wb = Workbook()
	ws = wb.active
	ws.title = "Employee Statistics"

	title = "تقرير إحصاءات الموظفين (Employee Statistics)"
	subtitle = (f"{summary['company_name']} | {summary['month_name_ar']} {summary['year']}"
	            f" | Salary reference: {summary.get('salary_col','')} | Currency: {summary.get('currency','')}")
	_write_header(ws, title, subtitle, 8, style)

	# summary
	srow = 4
	stats = [
		("Saudi %", summary["saudi_pct"]),
		("Non-Saudi %", summary["nonsaudi_pct"]),
		("Saudi Males", summary["saudi_male"]),
		("Saudi Females", summary["saudi_female"]),
		("Non-Saudi Males", summary["nonsaudi_male"]),
		("Non-Saudi Females", summary["nonsaudi_female"]),
		("Total Saudi Salaries", flt(summary["saudi_salaries"])),
		("Total Non-Saudi Salaries", flt(summary["nonsaudi_salaries"])),
	]
	for i, (k, v) in enumerate(stats):
		c = ws.cell(row=srow, column=1 + i, value=k)
		c.font = Font(bold=True, size=10)
		c2 = ws.cell(row=srow + 1, column=1 + i, value=flt(v))
		c2.font = Font(bold=True, size=10, color=_TEAL)
		c2.alignment = style["center"]
	srow += 3

	headers = ["Row #", "Employee ID", "Employee Name", "National ID / Iqama", "Nationality", "Gender", "Category", "Monthly Salary"]
	r = _write_table_header(ws, srow, headers, style)


	def _section(label, rows):
		nonlocal r
		ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)
		c = ws.cell(row=r, column=1, value=f"{label} ({len(rows)})")
		c.fill = style["sub_fill"]
		c.font = style["sub_font"]
		r += 1
		section_total = 0.0
		for idx, row in enumerate(rows, start=1):
			section_total += row["monthly_salary"]
			gender = "Male" if row["gender"] == "Male" else ("Female" if row["gender"] == "Female" else "-")
			vals = [idx, row["employee"], row["employee_name"], row["national_id"] or "-",
			        row["nationality"] or "-", gender, row["category"], flt(row["monthly_salary"])]
			for col, v in enumerate(vals, start=1):
				c = ws.cell(row=r, column=col, value=v)
				c.border = style["border"]
				c.alignment = style["right"] if col in (2, 3, 5) else style["center"]
			r += 1
		ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
		c = ws.cell(row=r, column=1, value="Sub-total")
		c.fill = style["total_fill"]
		c.font = style["total_font"]
		c2 = ws.cell(row=r, column=8, value=flt(section_total))
		c2.fill = style["total_fill"]
		c2.font = style["total_font"]
		r += 1


	_section("Saudi Employees (سعودي)", saudi_rows)
	_section("Non-Saudi Employees (غير سعودي)", nonsaudi_rows)

	# overall total
	ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
	c = ws.cell(row=r, column=1, value="Grand Total")
	c.fill = style["total_fill"]
	c.font = Font(bold=True, size=12, color="FFFFFF")
	c2 = ws.cell(row=r, column=8, value=flt(summary["total_salaries"]))
	c2.fill = style["total_fill"]
	c2.font = Font(bold=True, size=12, color="FFFFFF")

	_autofit(ws, [8, 14, 28, 20, 16, 10, 16, 16])
	ws.freeze_panes = f"A{srow+2}"
	filename = f"Employee_Statistics_{summary['month']:02d}_{summary['year']}.xlsx"
	return _save_xlsx(wb, filename, "Employees", summary["month"], summary["year"], summary["company"])
